"""
ContextOS ingestors/csv_excel.py — CSV and Excel tabular data ingestion.

Converts CSV/XLSX/XLS to Markdown tables. Each sheet becomes a section.
Uses stdlib csv for CSV files. Uses openpyxl for Excel (optional).
Falls back gracefully if openpyxl not installed.
"""
from __future__ import annotations
import csv
import time
from pathlib import Path
from contextos.ingestors import register


def _csv_to_md_table(rows: list[list[str]], max_rows: int = 50) -> str:
    if not rows:
        return ""
    header = rows[0]
    lines  = ["| " + " | ".join(str(c) for c in header) + " |",
              "|" + "|".join(["---"] * len(header)) + "|"]
    for row in rows[1:max_rows + 1]:
        # Pad or trim row to header length
        padded = list(row) + [""] * max(0, len(header) - len(row))
        lines.append("| " + " | ".join(str(c)[:50] for c in padded[:len(header)]) + " |")
    if len(rows) > max_rows + 1:
        lines.append(f"\n*… {len(rows) - max_rows - 1} more rows not shown*")
    return "\n".join(lines)


def _frontmatter(path: Path, sheet_count: int) -> str:
    return (
        "---\n"
        "project: unknown\n"
        "type: note\n"
        "status: draft\n"
        "source: tabular\n"
        f"updated_at: {time.strftime('%Y-%m-%d')}\n"
        "tags:\n"
        f"  - {path.suffix.lstrip('.')}\n"
        "  - data\n"
        f"sheets: {sheet_count}\n"
        "---\n"
    )


@register(".csv", "stdlib")
def extract_csv(path: Path) -> str:
    try:
        with open(path, encoding="utf-8-sig", errors="ignore", newline="") as f:
            rows = list(csv.reader(f))
    except Exception as exc:
        return f"---\nproject: unknown\ntype: note\nstatus: draft\n---\n# {path.name}\n\nCould not read: {exc}\n"

    fm      = _frontmatter(path, 1)
    title   = path.stem.replace("-", " ").replace("_", " ").title()
    table   = _csv_to_md_table(rows)
    return f"{fm}\n# {title}\n\n**Rows:** {max(0, len(rows)-1)}  **Columns:** {len(rows[0]) if rows else 0}\n\n{table}\n"


@register(".xlsx", "openpyxl")
@register(".xls",  "openpyxl")
def extract_excel(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except ImportError:
        return (
            f"---\nproject: unknown\ntype: note\nstatus: draft\n---\n"
            f"# {path.name}\n\nopenpyxl not installed. Run: pip install openpyxl\n"
        )
    except Exception as exc:
        return f"---\nproject: unknown\ntype: note\nstatus: draft\n---\n# {path.name}\n\nCould not read: {exc}\n"

    fm     = _frontmatter(path, len(wb.sheetnames))
    title  = path.stem.replace("-", " ").replace("_", " ").title()
    sections = [f"{fm}\n# {title}\n\n**Sheets:** {len(wb.sheetnames)}\n"]

    for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
        ws   = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(max_row=52, values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
            if len(rows) > 51:
                break
        if rows:
            sections.append(f"## {sheet_name}\n\n{_csv_to_md_table(rows)}")

    return "\n\n".join(sections)
